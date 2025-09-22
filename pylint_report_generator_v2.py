import re
import subprocess
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from utils import Utils

config = Utils.get_config()


def run_pylint_and_export(file_paths, excel_file=f"{config['pylint_report_path']}Pylint_report.xlsx"):
    """
    Run pylint on given files and export results into an Excel report.

    Args:
        file_paths (list[str]): List of file paths or modules to analyze.
        excel_file (str): Output Excel filename.
    """

    # Run pylint as subprocess
    result = subprocess.run(
        ["pylint", "--output-format=text"] + file_paths,
        capture_output=True,
        text=True
    )

    output = result.stdout

    # Regex to capture pylint issues
    issue_pattern = re.compile(
        r"^(?P<filename>[\w\\\/\.\-]+):(?P<line>\d+):\d+: [A-Z]\d+: (?P<msg>.+)$",
        re.MULTILINE,
    )

    issues = []
    for match in issue_pattern.finditer(output):
        issues.append({
            "Filename": match.group("filename").split("\\")[-1].split("/")[-1],
            "Line number": int(match.group("line")),
            "Issue description": match.group("msg")
        })

    # Extract scores
    score_pattern = re.compile(
        r"Your code has been rated at ([\d\.]+)/10 \(previous run: ([\d\.]+)/10"
    )
    score_match = score_pattern.search(output)
    current_score, prev_score = None, None
    if score_match:
        current_score, prev_score = score_match.groups()

    # Write issues to Excel
    df = pd.DataFrame(issues)
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pylint Report")

    # Append scores manually with formatting
    wb = load_workbook(excel_file)
    ws = wb["Pylint Report"]
    last_row = ws.max_row + 2

    ws.cell(row=last_row, column=1, value="Current Score:")
    ws.cell(row=last_row, column=2, value=current_score)
    ws.cell(row=last_row, column=2).font = Font(bold=True)

    ws.cell(row=last_row + 1, column=1, value="Previous Score:")
    ws.cell(row=last_row + 1, column=2, value=prev_score)

    wb.save(excel_file)


if __name__ == "__main__":
    # Example: analyze your whole project
    run_pylint_and_export(["."])
